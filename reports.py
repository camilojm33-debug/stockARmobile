"""Reportes CSV, Excel y PDF para ventas, compras, gastos y balance."""

import csv
from datetime import date
from io import BytesIO, StringIO

from flask import Blueprint, make_response, render_template, request, send_file
from flask_login import current_user
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app import tenant_required, utcnow
from stockarmobile.helpers.dates import local_day_bounds_utc_naive, local_today, format_local_datetime, parse_date_yyyy_mm_dd

bp = Blueprint("reports", __name__)


def _company_timezone_name():
    from app import Company
    company = Company.query.filter_by(id=getattr(current_user, "company_id", None)).first()
    return (getattr(company, "timezone", None) or "America/Argentina/Buenos_Aires").strip() or "America/Argentina/Buenos_Aires"


def _parse_requested_day(value, fallback=None):
    parsed = parse_date_yyyy_mm_dd(value)
    if parsed is not None:
        return parsed.date()
    return fallback


def _local_date_range(start_value=None, end_value=None):
    tz_name = _company_timezone_name()
    today = local_today(tz_name)
    start_day = _parse_requested_day(start_value, today if not start_value else None)
    end_day = _parse_requested_day(end_value, today if not end_value else None)
    if start_day is None and end_day is None:
        start_day = end_day = today
    elif start_day is None:
        start_day = end_day
    elif end_day is None:
        end_day = start_day
    start_utc, _ = local_day_bounds_utc_naive(start_day, tz_name)
    _, end_utc = local_day_bounds_utc_naive(end_day, tz_name)
    return start_utc, end_utc, start_day, end_day, tz_name


@bp.route("/")
@tenant_required
def index():
    from app import Expense, PurchaseOrder, Sale, db, scope_query_to_company
    today = local_today(_company_timezone_name())
    desde_raw = (request.args.get("desde") or today.isoformat()).strip()
    hasta_raw = (request.args.get("hasta") or today.isoformat()).strip()
    start_utc, end_utc, start_day, end_day, tz_name = _local_date_range(desde_raw, hasta_raw)
    def _apply_date(q, col):
        return q.filter(col >= start_utc, col < end_utc)
    sales_total = _apply_date(scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(Sale.total_amount), 0)), Sale), Sale.date).scalar() or 0
    purchases_total = _apply_date(scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(PurchaseOrder.total_amount), 0)), PurchaseOrder), PurchaseOrder.date).scalar() or 0
    expenses_total = _apply_date(scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)), Expense), Expense.date).scalar() or 0
    return render_template("reportes/index.html", sales_total=sales_total, purchases_total=purchases_total, expenses_total=expenses_total, desde=start_day.isoformat(), hasta=end_day.isoformat(), report_timezone=tz_name)


@bp.route("/balance.csv")
@tenant_required
def balance_csv():
    from app import Expense, PurchaseOrder, Sale, db, scope_query_to_company
    rows = [
        ("Ventas", scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(Sale.total_amount), 0)), Sale).scalar() or 0),
        ("Compras", scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(PurchaseOrder.total_amount), 0)), PurchaseOrder).scalar() or 0),
        ("Gastos", scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)), Expense).scalar() or 0),
    ]
    output = StringIO(); writer = csv.writer(output); writer.writerow(["Concepto", "Importe"]); writer.writerows(rows)
    response = make_response(output.getvalue()); response.headers["Content-Type"] = "text/csv; charset=utf-8"; response.headers["Content-Disposition"] = f'attachment; filename="balance_{local_today(_company_timezone_name()):%Y%m%d}.csv"'; return response


@bp.route("/<kind>.csv")
@tenant_required
def export_csv(kind):
    rows, filename = _rows_for(kind); output = StringIO(); writer = csv.writer(output); writer.writerows(rows)
    response = make_response(output.getvalue()); response.headers["Content-Type"] = "text/csv; charset=utf-8"; response.headers["Content-Disposition"] = f'attachment; filename="{filename}_{local_today(_company_timezone_name()):%Y%m%d}.csv"'; return response


@bp.route("/<kind>.xlsx")
@tenant_required
def export_excel(kind):
    rows, filename = _rows_for(kind); workbook = Workbook(); sheet = workbook.active; sheet.title = filename[:31]
    header_fill = PatternFill("solid", fgColor="2563EB"); header_font = Font(color="FFFFFF", bold=True)
    for row_index, row in enumerate(rows, start=1):
        sheet.append(row)
        if row_index == 1:
            for cell in sheet[row_index]: cell.fill = header_fill; cell.font = header_font
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2; sheet.column_dimensions[column[0].column_letter].width = min(width, 42)
    buffer = BytesIO(); workbook.save(buffer); buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"{filename}_{local_today(_company_timezone_name()):%Y%m%d}.xlsx")


@bp.route("/balance.pdf")
@tenant_required
def balance_pdf():
    from app import Expense, PurchaseOrder, Sale, db, scope_query_to_company
    sales_total = scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(Sale.total_amount), 0)), Sale).scalar() or 0
    purchases_total = scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(PurchaseOrder.total_amount), 0)), PurchaseOrder).scalar() or 0
    expenses_total = scope_query_to_company(db.session.query(db.func.coalesce(db.func.sum(Expense.amount), 0)), Expense).scalar() or 0
    profit = sales_total - purchases_total - expenses_total
    buffer = BytesIO(); pdf = canvas.Canvas(buffer, pagesize=letter); pdf.setTitle("Balance StockArmobile")
    pdf.drawString(72, 740, "StockArmobile - Balance"); pdf.drawString(72, 700, f"Ventas: ${sales_total:.2f}"); pdf.drawString(72, 680, f"Compras: ${purchases_total:.2f}"); pdf.drawString(72, 660, f"Gastos: ${expenses_total:.2f}"); pdf.drawString(72, 630, f"Resultado: ${profit:.2f}"); pdf.save(); buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name="balance.pdf")


def _rows_for(kind):
    from app import CashMovement, Client, Expense, Product, PurchaseOrder, Sale, SaleItem, scope_query_to_company
    from sqlalchemy.orm import selectinload
    start_utc, end_utc, _, _, tz_name = _local_date_range(request.args.get("desde"), request.args.get("hasta"))
    if kind == "ventas":
        query = _date_filter(scope_query_to_company(Sale.query.options(selectinload(Sale.items).selectinload(SaleItem.product)).order_by(Sale.date.desc()), Sale), Sale.date, start_utc, end_utc)
        return [["ID", "Cliente", "Total", "Fecha", "Unidades"]] + [[s.id, s.customer or "", s.total_amount or 0, format_local_datetime(s.date, tz_name), "; ".join(f"{item.product.name if item.product else f'Producto {item.product_id}'} ({item.product.unit_measure if item.product else 'u'}) x {item.quantity:g}" for item in s.items)] for s in query.all()], "ventas"
    if kind == "compras":
        query = _date_filter(scope_query_to_company(PurchaseOrder.query.order_by(PurchaseOrder.date.desc()), PurchaseOrder), PurchaseOrder.date, start_utc, end_utc)
        return [["ID", "Proveedor", "Total", "Fecha"]] + [[p.id, p.supplier.name if p.supplier else "", p.total_amount or 0, format_local_datetime(p.date, tz_name)] for p in query.all()], "compras"
    if kind == "gastos":
        query = _date_filter(scope_query_to_company(Expense.query.order_by(Expense.date.desc()), Expense), Expense.date, start_utc, end_utc)
        return [["ID", "Categoria", "Descripcion", "Importe", "Fecha"]] + [[e.id, e.category, e.description, e.amount, format_local_datetime(e.date, tz_name)] for e in query.all()], "gastos"
    if kind == "caja":
        query = _date_filter(scope_query_to_company(CashMovement.query.order_by(CashMovement.created_at.desc()), CashMovement), CashMovement.created_at, start_utc, end_utc)
        return [["ID", "Tipo", "Categoria", "Importe", "Fecha"]] + [[m.id, m.movement_type, m.category or "", m.amount, format_local_datetime(m.created_at, tz_name)] for m in query.all()], "caja"
    if kind == "clientes": return [["ID", "Nombre", "Email", "WhatsApp", "Saldo", "Credito"]] + [[c.id, c.name, c.email or "", c.whatsapp or "", c.balance or 0, c.credit_limit or 0] for c in scope_query_to_company(Client.query.order_by(Client.name), Client).all()], "clientes"
    if kind == "productos": return [["ID", "Codigo", "Nombre", "Marca", "Unidad", "Stock", "Precio", "Costo"]] + [[p.id, p.barcode, p.name, p.brand or "", p.unit_measure or "", p.stock or 0, p.price or 0, p.cost_price or 0] for p in scope_query_to_company(Product.query.order_by(Product.name), Product).all()], "productos"
    if kind == "stock": return [["Codigo", "Producto", "Stock", "Minimo", "Unidad", "Estado"]] + [[p.barcode, p.name, p.stock or 0, p.min_stock or 0, p.unit_measure or "", "critico" if (p.stock or 0) <= (p.min_stock or 0) else "ok"] for p in scope_query_to_company(Product.query.order_by(Product.name), Product).all()], "stock"
    return [["Error"], ["Reporte no reconocido"]], "reporte"


def _date_filter(query, column, start, end):
    if start is not None: query = query.filter(column >= start)
    if end is not None: query = query.filter(column < end)
    return query
