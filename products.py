"""Blueprint de productos: CRUD e inventario."""

import os
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import Blueprint, current_app, flash, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from sqlalchemy.exc import IntegrityError
from app import tenant_required, utcnow

bp = Blueprint("products", __name__)

ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_IMAGE_SIZE_BYTES = 5 * 1024 * 1024


def _product_to_dict(product):
    return {
        "id": product.id,
        "barcode": product.barcode,
        "codigo": product.barcode,
        "name": product.name,
        "nombre": product.name,
        "description": product.description or "",
        "category": product.category or "",
        "categoria": product.category or "",
        "sale_type": product.sale_type or "unidad",
        "tipo_venta": product.sale_type or "unidad",
        "unit_measure": product.unit_measure or "u",
        "unidad_medida": product.unit_measure or "u",
        "photo": product.photo or "",
        "brand": product.brand or "",
        "marca": product.brand or "",
        "supplier": product.supplier or "",
        "proveedor": product.supplier or "",
        "cost_price": float(product.cost_price or 0),
        "precio_costo": float(product.cost_price or 0),
        "price": float(product.price or 0),
        "precio_venta": float(product.price or 0),
        "margin": float(product.margin or 0),
        "profit_percent": float(product.profit_percent or 0),
        "tax": float(product.tax or 0),
        "iva": float(product.tax or 0),
        "stock": product.stock or 0,
        "min_stock": product.min_stock or 0,
        "discount": float(product.discount or 0),
        "favorite": bool(product.favorite),
    }


def _apply_product_form(product, form):
    product.barcode = (form.barcode.data or product.barcode or "").strip()
    product.name = form.name.data
    product.description = form.description.data
    product.category = form.category.data
    product.sale_type = form.sale_type.data or "unidad"
    product.unit_measure = (form.unit_measure.data or "").strip() or _default_unit(product.sale_type)
    product.brand = form.brand.data
    product.supplier = form.supplier.data
    cost_price = float(form.cost_price.data or 0)
    tax_percent = float(form.tax.data or 0)
    raw_price = (request.form.get("price") or "").strip()
    raw_profit_percent = (request.form.get("profit_percent") or "").strip()
    raw_margin = (request.form.get("margin") or "").strip()
    pricing_source = (request.form.get("pricing_source") or "").strip().lower()

    if pricing_source not in {"price", "profit_percent", "margin"}:
        if request.endpoint == "products.edit":
            posted_price = float(form.price.data or 0)
            posted_margin = float(form.margin.data or 0)
            posted_profit = float(form.profit_percent.data or 0)
            current_price = float(product.price or 0)
            current_margin = float(product.margin or 0)
            current_profit = float(product.profit_percent or 0)
            if raw_margin and posted_margin != current_margin:
                pricing_source = "margin"
            elif raw_profit_percent and posted_profit != current_profit:
                pricing_source = "profit_percent"
            elif raw_price and posted_price != current_price:
                pricing_source = "price"

    if pricing_source == "profit_percent" and raw_profit_percent:
        margin_percent = float(form.profit_percent.data or 0)
        if margin_percent < 0:
            raise ValueError("El margen % no puede ser negativo.")
        final_price = cost_price * (1 + (margin_percent / 100))
        gain_amount = final_price - cost_price
    elif pricing_source == "margin" and raw_margin:
        gain_amount = float(form.margin.data or 0)
        if gain_amount < 0:
            raise ValueError("La ganancia no puede ser negativa.")
        final_price = cost_price + gain_amount
        margin_percent = (gain_amount / cost_price * 100) if cost_price > 0 else 0.0
    elif pricing_source == "price" and raw_price:
        final_price = float(form.price.data or 0)
        if final_price < 0:
            raise ValueError("El precio de venta no puede ser negativo.")
        gain_amount = final_price - cost_price
        if gain_amount < 0:
            raise ValueError("El precio de venta no puede ser menor al costo.")
        margin_percent = (gain_amount / cost_price * 100) if cost_price > 0 else 0.0
    elif raw_profit_percent:
        margin_percent = float(form.profit_percent.data or 0)
        if margin_percent < 0:
            raise ValueError("El margen % no puede ser negativo.")
        final_price = cost_price * (1 + (margin_percent / 100))
        gain_amount = final_price - cost_price
    elif raw_margin:
        gain_amount = float(form.margin.data or 0)
        if gain_amount < 0:
            raise ValueError("La ganancia no puede ser negativa.")
        final_price = cost_price + gain_amount
        margin_percent = (gain_amount / cost_price * 100) if cost_price > 0 else 0.0
    elif raw_price:
        final_price = float(form.price.data or 0)
        if final_price < 0:
            raise ValueError("El precio de venta no puede ser negativo.")
        gain_amount = final_price - cost_price
        if gain_amount < 0:
            raise ValueError("El precio de venta no puede ser menor al costo.")
        margin_percent = (gain_amount / cost_price * 100) if cost_price > 0 else 0.0
    else:
        # Compatibilidad con datos existentes y formularios parciales.
        final_price = float(form.price.data or product.price or 0)
        if final_price < 0:
            raise ValueError("El precio de venta no puede ser negativo.")
        gain_amount = final_price - cost_price
        if gain_amount < 0:
            raise ValueError("El precio de venta no puede ser menor al costo.")
        margin_percent = (gain_amount / cost_price * 100) if cost_price > 0 else 0.0

    product.cost_price = cost_price
    product.price = final_price
    product.margin = gain_amount
    product.profit_percent = margin_percent
    product.tax = tax_percent
    product.stock = float(form.stock.data or 0)
    product.min_stock = float(form.min_stock.data or 0)
    product.discount = float(form.discount.data or 0)
    product.favorite = bool(form.favorite.data)


def _default_unit(sale_type):
    return {
        "unidad": "u",
        "kilogramo": "kg",
        "gramos": "g",
        "litros": "l",
        "mililitros": "ml",
        "metros": "m",
    }.get(sale_type or "unidad", "u")


def _float_value(value, default=0.0):
    try:
        return float(value if value not in (None, "") else default)
    except (TypeError, ValueError):
        return default


def _save_product_image(upload):
    filename = (upload.filename or "").strip()
    if not filename:
        return None

    extension = Path(filename).suffix.lower()
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Formato de imagen no permitido. Usa JPG, JPEG, PNG o WEBP.")

    upload.stream.seek(0, os.SEEK_END)
    size = upload.stream.tell()
    upload.stream.seek(0)
    if size > MAX_IMAGE_SIZE_BYTES:
        raise ValueError("La imagen supera el tamaño máximo de 5 MB.")

    upload_dir = Path(current_app.static_folder) / "uploads" / "products"
    upload_dir.mkdir(parents=True, exist_ok=True)
    unique_name = f"{uuid.uuid4().hex}{extension}"
    destination = upload_dir / unique_name
    upload.save(destination)
    return f"/static/uploads/products/{unique_name}"


@bp.route("/")
@tenant_required
def index():
    from app import Product, ProductForm, scope_query_to_company

    query = scope_query_to_company(Product.query.filter_by(active=True), Product)
    search = request.args.get("q") or request.args.get("search")
    category = request.args.get("categoria") or request.args.get("category")
    low_stock = request.args.get("low_stock")
    if search:
        like = f"%{search}%"
        query = query.filter((Product.name.ilike(like)) | (Product.barcode.ilike(like)) | (Product.category.ilike(like)))
    if category:
        query = query.filter(Product.category == category)
    if low_stock:
        query = query.filter(Product.stock <= Product.min_stock)

    productos = query.order_by(Product.name).all()
    categorias = [c[0] for c in scope_query_to_company(Product.query.with_entities(Product.category), Product).distinct().all() if c[0]]
    return render_template(
        "productos/index.html",
        products=productos,
        productos=productos,
        categorias=categorias,
        form=ProductForm(),
        edit=False,
    )


@bp.route("/add", methods=["GET", "POST"])
@tenant_required
def add():
    from app import Product, ProductForm, ProductModification, db, scope_query_to_company

    form = ProductForm()
    if form.validate_on_submit():
        next_id = (db.session.query(db.func.coalesce(db.func.max(Product.id), 0)).scalar() or 0) + 1
        barcode = (form.barcode.data or "").strip() or f"P{next_id:06d}"
        if scope_query_to_company(Product.query.filter_by(barcode=barcode), Product).first():
            flash("El codigo de barras ya existe.", "danger")
            return redirect(url_for("products.index"))

        product = Product(barcode=barcode, name=form.name.data, company_id=getattr(current_user, 'company_id', None))
        try:
            _apply_product_form(product, form)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("products.index"))
        upload = request.files.get("photo_file")
        if upload and (upload.filename or "").strip():
            try:
                product.photo = _save_product_image(upload)
            except ValueError as exc:
                flash(str(exc), "danger")
                return redirect(url_for("products.index"))
        try:
            db.session.add(product)
            db.session.flush()
            db.session.add(
                ProductModification(
                    product_id=product.id,
                    company_id=product.company_id,
                    user_id=current_user.id,
                    action="creacion",
                    detail="Producto creado",
                )
            )
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("No se pudo guardar: el codigo de barras ya existe.", "danger")
            return redirect(url_for("products.index"))
        flash("Producto creado exitosamente.", "success")
        return redirect(url_for("products.index"))

    return render_template("productos/index.html", productos=[], products=[], categorias=[], form=form, edit=True)


@bp.route("/edit/<int:product_id>", methods=["GET", "POST"])
@bp.route("/edit/<int:id>", methods=["GET", "POST"])
@tenant_required
def edit(product_id=None, id=None):
    from app import Product, ProductForm, ProductModification, ProductPriceHistory, db, scope_query_to_company

    product = scope_query_to_company(db.session.query(Product), Product).filter(Product.id == (product_id or id)).first()
    if product is None:
        flash("Producto no encontrado.", "warning")
        return redirect(url_for("products.index"))

    form = ProductForm(obj=product)
    if form.validate_on_submit():
        old_price = float(product.price or 0)
        old_cost = float(product.cost_price or 0)
        try:
            _apply_product_form(product, form)
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("products.edit", product_id=product.id))
        upload = request.files.get("photo_file")
        if upload and (upload.filename or "").strip():
            try:
                product.photo = _save_product_image(upload)
            except ValueError as exc:
                flash(str(exc), "danger")
                return redirect(url_for("products.edit", product_id=product.id))
        if old_price != float(product.price or 0) or old_cost != float(product.cost_price or 0):
            db.session.add(
                ProductPriceHistory(
                    product_id=product.id,
                    company_id=product.company_id,
                    user_id=current_user.id,
                    old_price=old_price,
                    new_price=float(product.price or 0),
                    old_cost=old_cost,
                    new_cost=float(product.cost_price or 0),
                )
            )
        db.session.add(
            ProductModification(
                product_id=product.id,
                company_id=product.company_id,
                user_id=current_user.id,
                action="edicion",
                detail="Producto actualizado",
            )
        )
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            flash("No se pudo actualizar: el codigo de barras ya existe.", "danger")
            return redirect(url_for("products.index"))
        flash("Producto actualizado exitosamente.", "success")
        return redirect(url_for("products.index"))

    categorias = [c[0] for c in scope_query_to_company(Product.query.with_entities(Product.category), Product).distinct().all() if c[0]]
    return render_template("productos/index.html", productos=[product], products=[product], categorias=categorias, form=form, edit=True)


@bp.route("/delete/<int:product_id>", methods=["POST"])
@bp.route("/delete/<int:id>", methods=["POST"])
@tenant_required
def delete(product_id=None, id=None):
    from app import Product, db, scope_query_to_company

    product = scope_query_to_company(db.session.query(Product), Product).filter(Product.id == (product_id or id)).first()
    if product:
        product.active = False
        db.session.commit()
        flash("Producto desactivado exitosamente.", "success")
    return redirect(url_for("products.index"))


@bp.route("/<int:product_id>/kardex")
@tenant_required
def kardex(product_id):
    from app import Product, ProductModification, PurchaseItem, PurchaseOrder, Sale, SaleItem, db, scope_query_to_company

    product = scope_query_to_company(db.session.query(Product), Product).filter(Product.id == product_id).first()
    if product is None:
        flash("Producto no encontrado.", "warning")
        return redirect(url_for("products.index"))

    movements = []
    sale_rows = (
        scope_query_to_company(
            db.session.query(Sale, SaleItem).join(SaleItem, Sale.id == SaleItem.sale_id),
            Sale,
        )
        .filter(SaleItem.product_id == product.id)
        .order_by(Sale.date.desc())
        .limit(100)
        .all()
    )
    for sale, item in sale_rows:
        movements.append(
            {
                "date": sale.date,
                "type": "Venta",
                "detail": f"Venta #{sale.id} - {sale.customer or 'Consumidor final'}",
                "quantity": -float(item.quantity or 0),
                "unit_value": float(item.price or 0),
            }
        )

    purchase_rows = (
        scope_query_to_company(
            db.session.query(PurchaseOrder, PurchaseItem).join(PurchaseItem, PurchaseOrder.id == PurchaseItem.purchase_order_id),
            PurchaseOrder,
        )
        .filter(PurchaseItem.product_id == product.id)
        .order_by(PurchaseOrder.date.desc())
        .limit(100)
        .all()
    )
    for purchase, item in purchase_rows:
        movements.append(
            {
                "date": purchase.date,
                "type": "Compra",
                "detail": f"Compra #{purchase.id}",
                "quantity": float(item.quantity or 0),
                "unit_value": float(item.unit_cost or 0),
            }
        )

    modifications = (
        scope_query_to_company(ProductModification.query.filter_by(product_id=product.id), ProductModification)
        .order_by(ProductModification.created_at.desc())
        .limit(50)
        .all()
    )
    for modification in modifications:
        movements.append(
            {
                "date": modification.created_at,
                "type": modification.action.capitalize(),
                "detail": modification.detail or "",
                "quantity": None,
                "unit_value": None,
            }
        )

    movements.sort(key=lambda item: item["date"] or datetime.min, reverse=True)
    return render_template("productos/kardex.html", product=product, movements=movements[:150])


@bp.route("/export.xlsx")
@tenant_required
def export_excel():
    from app import Product, scope_query_to_company
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"
    headers = ["barcode", "name", "category", "brand", "supplier", "cost_price", "price", "stock", "min_stock", "sale_type", "unit_measure", "discount"]
    sheet.append(headers)
    for product in scope_query_to_company(Product.query.filter_by(active=True), Product).order_by(Product.name).all():
        sheet.append(
            [
                product.barcode,
                product.name,
                product.category or "",
                product.brand or "",
                product.supplier or "",
                float(product.cost_price or 0),
                float(product.price or 0),
                float(product.stock or 0),
                float(product.min_stock or 0),
                product.sale_type or "unidad",
                product.unit_measure or "u",
                float(product.discount or 0),
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"productos_{utcnow():%Y%m%d}.xlsx",
    )


@bp.route("/import", methods=["POST"])
@tenant_required
def import_excel():
    from app import Product, ProductModification, db, scope_query_to_company
    from openpyxl import load_workbook

    upload = request.files.get("file")
    if not upload or not upload.filename.lower().endswith(".xlsx"):
        flash("Subi un archivo .xlsx valido.", "danger")
        return redirect(url_for("products.index"))

    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        flash("El archivo no contiene productos.", "warning")
        return redirect(url_for("products.index"))

    headers = [str(cell or "").strip().lower() for cell in rows[0]]
    created = 0
    updated = 0
    for row in rows[1:]:
        data = dict(zip(headers, row))
        barcode = str(data.get("barcode") or data.get("codigo") or "").strip()
        name = str(data.get("name") or data.get("nombre") or "").strip()
        if not barcode or not name:
            continue
        product = scope_query_to_company(Product.query.filter_by(barcode=barcode), Product).first()
        if product is None:
            product = Product(
                barcode=barcode,
                name=name,
                active=True,
                company_id=getattr(current_user, "company_id", None),
            )
            db.session.add(product)
            db.session.flush()
            created += 1
            action = "importacion"
        else:
            updated += 1
            action = "actualizacion_importacion"
        product.name = name
        product.active = True
        product.category = data.get("category") or data.get("categoria") or product.category
        product.brand = data.get("brand") or data.get("marca") or product.brand
        product.supplier = data.get("supplier") or data.get("proveedor") or product.supplier
        product.cost_price = _float_value(data.get("cost_price") or data.get("precio_costo"), product.cost_price or 0)
        product.price = _float_value(data.get("price") or data.get("precio_venta"), product.price or 0)
        product.stock = _float_value(data.get("stock"), product.stock or 0)
        product.min_stock = _float_value(data.get("min_stock") or data.get("stock_minimo"), product.min_stock or 0)
        product.sale_type = data.get("sale_type") or data.get("tipo_venta") or product.sale_type or "unidad"
        product.unit_measure = data.get("unit_measure") or data.get("unidad_medida") or product.unit_measure or _default_unit(product.sale_type)
        product.discount = _float_value(data.get("discount") or data.get("descuento"), product.discount or 0)
        product.margin = float(product.price or 0) - float(product.cost_price or 0)
        product.profit_percent = (product.margin / float(product.cost_price or 1)) * 100 if product.cost_price else 0
        db.session.add(
            ProductModification(
                product_id=product.id,
                company_id=product.company_id,
                user_id=current_user.id,
                action=action,
                detail="Importacion Excel",
            )
        )
    db.session.commit()
    flash(f"Importacion completada: {created} creados, {updated} actualizados.", "success")
    return redirect(url_for("products.index"))


@bp.route("/api/products")
@tenant_required
def api_list():
    from app import Product, scope_query_to_company

    return jsonify({"products": [_product_to_dict(p) for p in scope_query_to_company(Product.query.filter_by(active=True), Product).order_by(Product.name).all()]})


@bp.route("/api/<barcode>")
@tenant_required
def api_get_by_barcode(barcode):
    from app import Product, scope_query_to_company

    product = scope_query_to_company(Product.query.filter_by(barcode=barcode, active=True), Product).first()
    if not product:
        return jsonify({"error": "Producto no encontrado"}), 404
    return jsonify(_product_to_dict(product))
